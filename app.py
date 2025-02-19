from flask import Flask, request, send_file, jsonify
from flask_cors import CORS
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import os

app = Flask(__name__)
CORS(app)

# Function to process the Excel file
def process_excel(input_path, output_path):
    try:
        df = pd.read_excel(input_path, sheet_name="B2B", engine="openpyxl")
        
        # Check if the DataFrame is valid
        if df.empty:
            return False, "Error: Excel file is empty or sheet not found."

        # Ensure the specified indices exist before dropping
        rows_to_delete = [1, 2, 3, 5]
        rows_to_delete = [i for i in rows_to_delete if i < len(df)]
        df.drop(index=rows_to_delete, inplace=True)
        df.reset_index(drop=True, inplace=True)

        # Ensure the specified columns exist before dropping
        all_columns = list(df.columns)
        columns_to_delete = [3, 6, 7, 13, 14, 15, 16, 17, 18, 19, 20]
        columns_to_delete = [all_columns[i] for i in columns_to_delete if i < len(all_columns)]
        df.drop(columns=columns_to_delete, axis=1, inplace=True)

        # Modify 7th row headers (adjust for index)
        if len(df) > 1:
            df.iloc[1] = ["GSTIN", "LEGAL NAME", "INV NO", "INV DATE", "INV VALUE", "TAXABLE VALUE", "IGST", "CGST", "SGST", "CESS"]

        # Insert an empty first row
        df.loc[-1] = [""] * len(df.columns)
        df.index = df.index + 1
        df = df.sort_index()

        # Add "Total" row at the bottom
        df.loc[len(df)] = ["Total"] + [None] * (len(df.columns) - 1)

        # Save modified DataFrame to Excel
        df.to_excel(output_path, index=False, engine="openpyxl")

        # Open with openpyxl for formatting
        wb = openpyxl.load_workbook(output_path)
        ws = wb.active

        # Define styles
        bold_font = Font(bold=True, size=12)
        center_alignment = Alignment(horizontal="center")
        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        thin_border = Border(left=Side(style='thin'),
                             right=Side(style='thin'),
                             top=Side(style='thin'),
                             bottom=Side(style='thin'))

        # Apply Bold to Header Row
        if ws.max_row > 1:
            for cell in ws[4]:  # Header row now at index 7
                cell.font = bold_font
                cell.alignment = center_alignment
                cell.fill = yellow_fill

        # Apply SUM formula to the last row of numeric columns
        total_row_index = ws.max_row
        for col_idx in range(5, 11):  # Columns E to J (1-based index)
            col_letter = get_column_letter(col_idx)
            sum_formula = f"=SUM({col_letter}5:{col_letter}{total_row_index - 1})"
            ws[f"{col_letter}{total_row_index}"].value = sum_formula
            ws[f"{col_letter}{total_row_index}"].font = bold_font

        # Apply Bold to "Total" Row
        for cell in ws[total_row_index]:
            cell.font = bold_font

        # Set Column Widths based on content
        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            ws.column_dimensions[col_letter].width = max_length + 2

        # Apply Borders to All Cells
        for row in ws.iter_rows():
            for cell in row:
                cell.border = thin_border

        # Format Numeric Columns (Add thousand separator)
        for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                if isinstance(cell.value, (int, float)):
                    cell.number_format = "#,##0.00"

        wb.save(output_path)
        return True, "File processed successfully."
    except Exception as e:
        return False, str(e)

@app.route("/upload", methods=["POST"])
def upload_file():
    if "file" not in request.files:
        return jsonify({"error": "No file part"}), 400
    
    file = request.files["file"]
    if file.filename == "":
        return jsonify({"error": "No selected file"}), 400
    
    input_path = "temp_input.xlsx"
    output_path = "temp_output.xlsx"
    file.save(input_path)
    
    success, message = process_excel(input_path, output_path)
    if not success:
        return jsonify({"error": message}), 500
    
    return send_file(output_path, as_attachment=True, download_name="Formatted_GSTR2B.xlsx")

import os

if __name__ == "__main__":
    port = int(os.environ.get("PORT", 10000))  # Default to 10000 if PORT is not set
    app.run(host="0.0.0.0", port=port)
